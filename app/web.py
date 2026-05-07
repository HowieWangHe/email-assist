from __future__ import annotations

import csv
from datetime import datetime, timezone
from io import BytesIO, StringIO
import mimetypes
from pathlib import Path
import shutil
import subprocess
from uuid import uuid4

from fastapi import APIRouter, File, Form, Request, UploadFile, status
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from pydantic import BaseModel

from app.models import Campaign, CampaignAttachment, CampaignRecipient, RecipientStatus, ReminderStrategy
from app.services.connectivity import ConnectivityTester
from app.services.campaigns import CampaignService
from app.services.exporting import ExportService
from app.services.extraction import OpenAICompatibleClient
from app.services.mail import MailboxFetchService
from app.services.processing import AttachmentProcessingService
from app.services.receiving import ReplyIngestService
from app.services.reminders import ReminderService
from app.services.sending import CampaignSendService
from app.services.storage import CampaignStorage
from app.settings_store import AiProviderConfig, AppLocalConfig, MailServerConfig


router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


def _format_dt(value):
    if value is None:
        return "—"
    return value.strftime("%Y-%m-%d %H:%M:%S")


templates.env.filters["dt"] = _format_dt


class RecipientInput(BaseModel):
    email: str
    name: str | None = None
    company: str | None = None


class CampaignCreateInput(BaseModel):
    name: str
    subject: str
    body_template: str
    deadline: datetime
    recipients: list[RecipientInput]
    reminder_strategy: ReminderStrategy = ReminderStrategy.MANUAL_CONFIRM
    attachment_ai_enabled: bool = False


class MailServerInput(BaseModel):
    host: str = ""
    port: int = 0
    security: str = "ssl"
    username: str = ""
    password: str = ""


class AiProviderInput(BaseModel):
    base_url: str = ""
    api_key: str = ""
    model: str = ""


class AppSettingsInput(BaseModel):
    smtp: MailServerInput = MailServerInput()
    imap: MailServerInput = MailServerInput()
    ai: AiProviderInput = AiProviderInput()


class ClaimReviewMessageInput(BaseModel):
    recipient_id: str


class ConfirmActionInput(BaseModel):
    confirm: str = ""


@router.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@router.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    campaigns = request.app.state.database.list_campaigns()
    _sync_campaign_deadline_statuses(request, campaigns)
    campaigns = request.app.state.database.list_campaigns()
    archived_campaigns = request.app.state.database.list_archived_campaigns()
    reminder_counts = _reminder_counts(request, campaigns)
    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "campaigns": campaigns,
            "archived_campaigns": archived_campaigns,
            "reminder_counts": reminder_counts,
        },
    )


@router.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request) -> HTMLResponse:
    local_config = request.app.state.settings_store.load()
    return templates.TemplateResponse(
        request,
        "settings.html",
        {"settings": request.app.state.settings, "local_config": local_config.masked()},
    )


@router.get("/campaigns/new", response_class=HTMLResponse)
def new_campaign_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(request, "campaign_new.html", {})


@router.get("/campaigns/{campaign_id}", response_class=HTMLResponse)
def campaign_detail_page(campaign_id: str, request: Request) -> HTMLResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return HTMLResponse("Campaign not found", status_code=status.HTTP_404_NOT_FOUND)
    recipients = request.app.state.database.list_recipients(campaign_id)
    if not _is_archived(campaign):
        CampaignService().refresh_campaign_status(campaign, recipients, now=datetime.now(timezone.utc))
        request.app.state.database.save_recipients(recipients)
        request.app.state.database.save_campaign(campaign)
    sent_messages = request.app.state.database.list_sent_messages(campaign_id)
    attachments = request.app.state.database.list_campaign_attachments(campaign_id)
    received_messages = request.app.state.database.list_received_messages(campaign_id)
    received_attachments = request.app.state.database.list_received_attachments(campaign_id)
    extraction_records = request.app.state.database.list_extraction_records(campaign_id)
    review_messages = _prioritize_review_messages(
        request.app.state.database.list_review_messages(campaign_id),
        recipients,
    )
    due_reminders = [] if _is_archived(campaign) else ReminderService().due_recipients(
        campaign,
        recipients,
        datetime.now(timezone.utc),
    )
    return templates.TemplateResponse(
        request,
        "campaign_detail.html",
        {
            "campaign": campaign,
            "recipients": recipients,
            "sent_messages": sent_messages,
            "attachments": attachments,
            "received_messages": received_messages,
            "received_attachments": received_attachments,
            "extraction_records": extraction_records,
            "review_messages": review_messages,
            "due_reminders": due_reminders,
        },
    )


@router.post("/campaigns", response_class=HTMLResponse)
def create_campaign_form(
    request: Request,
    name: str = Form(...),
    subject: str = Form(...),
    body_template: str = Form(...),
    deadline: str = Form(...),
    recipients_text: str = Form(""),
    recipients_file: UploadFile | None = File(None),
    cc: str = Form(""),
    bcc: str = Form(""),
    attachments: list[UploadFile] = File(default=[]),
) -> HTMLResponse:
    recipient_rows = _dedupe_recipients(
        _parse_recipients_text(recipients_text) + _parse_recipients_file(recipients_file)
    )
    if not recipient_rows:
        return HTMLResponse("At least one recipient is required", status_code=status.HTTP_400_BAD_REQUEST)
    campaign, recipients = _create_campaign_objects(
        name=name,
        subject=subject,
        body_template=body_template,
        deadline=datetime.fromisoformat(deadline),
        recipients=[RecipientInput(**item) for item in recipient_rows],
        cc=_split_addresses(cc),
        bcc=_split_addresses(bcc),
    )
    request.app.state.database.save_campaign(campaign)
    request.app.state.database.save_recipients(recipients)
    saved_attachments = _save_uploaded_attachments(request, campaign.id, attachments)
    request.app.state.database.save_campaign_attachments(saved_attachments)
    return HTMLResponse(status_code=status.HTTP_303_SEE_OTHER, headers={"Location": f"/campaigns/{campaign.id}"})


@router.post("/campaigns/{campaign_id}/recipients", response_class=HTMLResponse)
def append_recipients_form(
    campaign_id: str,
    request: Request,
    recipients_text: str = Form(""),
    recipients_file: UploadFile | None = File(None),
) -> HTMLResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return HTMLResponse("Campaign not found", status_code=status.HTTP_404_NOT_FOUND)
    if _is_archived(campaign):
        return HTMLResponse("Archived campaign cannot be edited", status_code=status.HTTP_409_CONFLICT)

    rows = _dedupe_recipients(_parse_recipients_text(recipients_text) + _parse_recipients_file(recipients_file))
    existing = request.app.state.database.list_recipients(campaign_id)
    new_recipients = _new_recipients_for_campaign(campaign_id, rows, existing)
    if new_recipients:
        request.app.state.database.save_recipients(new_recipients)
        recipients = existing + new_recipients
        CampaignService().refresh_campaign_status(campaign, recipients, now=datetime.now(timezone.utc))
        request.app.state.database.save_campaign(campaign)
    return HTMLResponse(status_code=status.HTTP_303_SEE_OTHER, headers={"Location": f"/campaigns/{campaign_id}"})


@router.get("/api/settings")
def get_settings(request: Request) -> dict:
    return request.app.state.settings_store.load().masked()


@router.post("/api/settings")
def save_settings(payload: AppSettingsInput, request: Request) -> dict:
    existing = request.app.state.settings_store.load()
    config = _settings_from_payload(payload, existing)
    request.app.state.settings_store.save(config)
    return config.masked()


@router.post("/api/settings/test/smtp")
def test_smtp(payload: MailServerInput) -> dict:
    result = ConnectivityTester().test_smtp(_mail_config(payload))
    return {"ok": result.ok, "message": result.message}


@router.post("/api/settings/test/imap")
def test_imap(payload: MailServerInput) -> dict:
    result = ConnectivityTester().test_imap(_mail_config(payload))
    return {"ok": result.ok, "message": result.message}


@router.post("/api/settings/test/ai")
def test_ai(payload: AiProviderInput) -> dict:
    result = ConnectivityTester().test_ai(_ai_config(payload))
    return {"ok": result.ok, "message": result.message}


@router.get("/api/campaigns")
def list_campaigns(request: Request) -> list[dict]:
    campaigns = request.app.state.database.list_campaigns()
    _sync_campaign_deadline_statuses(request, campaigns)
    return [_campaign_payload(campaign) for campaign in request.app.state.database.list_campaigns()]


@router.post("/api/campaigns", status_code=status.HTTP_201_CREATED)
def create_campaign(payload: CampaignCreateInput, request: Request) -> JSONResponse:
    campaign, recipients = _create_campaign_objects(
        name=payload.name,
        subject=payload.subject,
        body_template=payload.body_template,
        deadline=payload.deadline,
        recipients=payload.recipients,
        reminder_strategy=payload.reminder_strategy,
        attachment_ai_enabled=payload.attachment_ai_enabled,
    )
    request.app.state.database.save_campaign(campaign)
    request.app.state.database.save_recipients(recipients)
    return JSONResponse(_campaign_payload(campaign), status_code=status.HTTP_201_CREATED)


@router.get("/api/campaigns/{campaign_id}")
def get_campaign(campaign_id: str, request: Request) -> JSONResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    recipients = request.app.state.database.list_recipients(campaign_id)
    if not _is_archived(campaign):
        CampaignService().refresh_campaign_status(campaign, recipients, now=datetime.now(timezone.utc))
        request.app.state.database.save_recipients(recipients)
        request.app.state.database.save_campaign(campaign)
    payload = _campaign_payload(campaign)
    payload["recipients"] = [_recipient_payload(recipient) for recipient in recipients]
    payload["received_messages"] = [
        _received_message_payload(message)
        for message in request.app.state.database.list_received_messages(campaign_id)
    ]
    return JSONResponse(payload)


@router.post("/api/campaigns/{campaign_id}/archive")
def archive_campaign(campaign_id: str, payload: ConfirmActionInput, request: Request) -> JSONResponse:
    if payload.confirm != "ARCHIVE":
        return JSONResponse({"error": "confirmation_required"}, status_code=status.HTTP_400_BAD_REQUEST)
    campaign = request.app.state.database.archive_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return JSONResponse(_campaign_payload(campaign))


@router.post("/api/campaigns/{campaign_id}/unarchive")
def unarchive_campaign(campaign_id: str, request: Request) -> JSONResponse:
    campaign = request.app.state.database.unarchive_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return JSONResponse(_campaign_payload(campaign))


@router.delete("/api/campaigns/{campaign_id}")
def delete_campaign(campaign_id: str, payload: ConfirmActionInput, request: Request) -> JSONResponse:
    if payload.confirm != "DELETE":
        return JSONResponse({"error": "confirmation_required"}, status_code=status.HTTP_400_BAD_REQUEST)
    deleted = request.app.state.database.delete_campaign(campaign_id)
    if not deleted:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    for path in (
        request.app.state.settings.data_dir / "campaigns" / campaign_id,
        request.app.state.settings.data_dir / "exports" / campaign_id,
    ):
        shutil.rmtree(path, ignore_errors=True)
    return JSONResponse({"campaign_id": campaign_id, "deleted": True})


@router.post("/api/campaigns/{campaign_id}/refresh")
def refresh_campaign(campaign_id: str, request: Request) -> JSONResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    if _is_archived(campaign):
        return _archived_response()
    config = request.app.state.settings_store.load()
    recipients = request.app.state.database.list_recipients(campaign_id)
    sent_messages = request.app.state.database.list_sent_messages(campaign_id)
    try:
        raw_messages = MailboxFetchService().fetch_recent(config.imap, limit=80)
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=status.HTTP_400_BAD_REQUEST)

    result = ReplyIngestService(request.app.state.settings.data_dir).ingest(
        raw_messages=raw_messages,
        campaigns=[campaign],
        recipients=recipients,
        sent_messages=sent_messages,
        known_message_ids=request.app.state.database.list_received_message_ids(),
    )
    request.app.state.database.save_recipients(recipients)
    request.app.state.database.save_received_messages(result.received_messages)
    request.app.state.database.save_received_attachments(result.attachments)
    CampaignService().refresh_campaign_status(campaign, recipients, now=datetime.now(timezone.utc))
    request.app.state.database.save_campaign(campaign)
    return JSONResponse(
        {
            "campaign_id": campaign_id,
            "fetched_count": len(raw_messages),
            "replied_count": result.replied_count,
            "review_count": result.review_count,
            "skipped_existing_count": result.skipped_existing_count,
            "attachment_count": len(result.attachments),
        }
    )


@router.post("/api/campaigns/{campaign_id}/review-messages/{message_row_id}/claim")
def claim_review_message(
    campaign_id: str,
    message_row_id: str,
    payload: ClaimReviewMessageInput,
    request: Request,
) -> JSONResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    if _is_archived(campaign):
        return _archived_response()
    recipients = request.app.state.database.list_recipients(campaign_id)
    recipient = next((item for item in recipients if item.id == payload.recipient_id), None)
    if recipient is None:
        return JSONResponse({"error": "recipient_not_found"}, status_code=status.HTTP_404_NOT_FOUND)

    claimed = request.app.state.database.claim_received_message(
        message_row_id,
        campaign_id=campaign_id,
        recipient_id=recipient.id,
    )
    if claimed is None:
        return JSONResponse({"error": "message_not_found"}, status_code=status.HTTP_404_NOT_FOUND)

    recipient.status = RecipientStatus.REPLIED
    recipient.replied_at = datetime.now()
    request.app.state.database.save_recipients([recipient])
    CampaignService().refresh_campaign_status(campaign, recipients)
    request.app.state.database.save_campaign(campaign)
    return JSONResponse(
        {
            "campaign_id": campaign_id,
            "message_id": claimed.message_id,
            "recipient_id": recipient.id,
            "recipient_status": recipient.status.value,
            "campaign_status": campaign.status,
        }
    )


@router.get("/api/campaigns/{campaign_id}/received-messages/{message_row_id}/body")
def preview_received_message_body(campaign_id: str, message_row_id: str, request: Request) -> JSONResponse:
    message = next(
        (
            item
            for item in request.app.state.database.list_received_messages(campaign_id)
            if item.id == message_row_id
        ),
        None,
    )
    if message is None:
        return JSONResponse({"error": "message_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    allowed_path = _safe_local_file(request, message.body_path)
    if allowed_path is None:
        return JSONResponse({"error": "file_not_allowed"}, status_code=status.HTTP_403_FORBIDDEN)
    if not allowed_path.exists():
        return JSONResponse({"error": "file_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return JSONResponse(
        {
            "id": message.id,
            "subject": message.subject,
            "from_email": message.from_email,
            "summary": message.body_summary,
            "body": allowed_path.read_text(encoding="utf-8", errors="replace"),
        }
    )


@router.get("/api/campaigns/{campaign_id}/received-attachments/{attachment_id}/preview")
def preview_received_attachment(campaign_id: str, attachment_id: str, request: Request) -> JSONResponse:
    attachment = _find_received_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _preview_file_response(request, attachment.path, attachment.content_type)


@router.get("/api/campaigns/{campaign_id}/campaign-attachments/{attachment_id}/preview")
def preview_campaign_attachment(campaign_id: str, attachment_id: str, request: Request) -> JSONResponse:
    attachment = _find_campaign_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _preview_file_response(request, attachment.path, attachment.content_type)


@router.get("/api/campaigns/{campaign_id}/received-attachments/{attachment_id}/view")
def view_received_attachment(campaign_id: str, attachment_id: str, request: Request):
    attachment = _find_received_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _inline_file_response(request, attachment.path, attachment.filename, attachment.content_type)


@router.get("/api/campaigns/{campaign_id}/campaign-attachments/{attachment_id}/view")
def view_campaign_attachment(campaign_id: str, attachment_id: str, request: Request):
    attachment = _find_campaign_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _inline_file_response(request, attachment.path, attachment.filename, attachment.content_type)


@router.get("/api/campaigns/{campaign_id}/received-attachments/{attachment_id}/download")
def download_received_attachment(campaign_id: str, attachment_id: str, request: Request):
    attachment = _find_received_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _download_file_response(request, attachment.path, attachment.filename, attachment.content_type)


@router.get("/api/campaigns/{campaign_id}/campaign-attachments/{attachment_id}/download")
def download_campaign_attachment(campaign_id: str, attachment_id: str, request: Request):
    attachment = _find_campaign_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _download_file_response(request, attachment.path, attachment.filename, attachment.content_type)


@router.post("/api/campaigns/{campaign_id}/received-messages/{message_row_id}/open")
def open_received_message_body(campaign_id: str, message_row_id: str, request: Request) -> JSONResponse:
    message = next(
        (
            item
            for item in request.app.state.database.list_received_messages(campaign_id)
            if item.id == message_row_id
        ),
        None,
    )
    if message is None:
        return JSONResponse({"error": "message_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _open_file_response(request, message.body_path)


@router.post("/api/campaigns/{campaign_id}/received-attachments/{attachment_id}/open")
def open_received_attachment(campaign_id: str, attachment_id: str, request: Request) -> JSONResponse:
    attachment = _find_received_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _open_file_response(request, attachment.path)


@router.post("/api/campaigns/{campaign_id}/campaign-attachments/{attachment_id}/open")
def open_campaign_attachment(campaign_id: str, attachment_id: str, request: Request) -> JSONResponse:
    attachment = _find_campaign_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _open_file_response(request, attachment.path)


@router.post("/api/campaigns/{campaign_id}/received-messages/{message_row_id}/reveal")
def reveal_received_message_body(campaign_id: str, message_row_id: str, request: Request) -> JSONResponse:
    message = next(
        (
            item
            for item in request.app.state.database.list_received_messages(campaign_id)
            if item.id == message_row_id
        ),
        None,
    )
    if message is None:
        return JSONResponse({"error": "message_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _reveal_file_response(request, message.body_path)


@router.post("/api/campaigns/{campaign_id}/received-attachments/{attachment_id}/reveal")
def reveal_received_attachment(campaign_id: str, attachment_id: str, request: Request) -> JSONResponse:
    attachment = _find_received_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _reveal_file_response(request, attachment.path)


@router.post("/api/campaigns/{campaign_id}/campaign-attachments/{attachment_id}/reveal")
def reveal_campaign_attachment(campaign_id: str, attachment_id: str, request: Request) -> JSONResponse:
    attachment = _find_campaign_attachment(request, campaign_id, attachment_id)
    if attachment is None:
        return JSONResponse({"error": "attachment_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    return _reveal_file_response(request, attachment.path)


@router.post("/api/campaigns/{campaign_id}/process-attachments")
def process_attachments(campaign_id: str, request: Request, use_ai: bool = False) -> JSONResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    if _is_archived(campaign):
        return _archived_response()
    attachments = request.app.state.database.list_received_attachments(campaign_id)
    ai_client = None
    if use_ai:
        config = request.app.state.settings_store.load().ai
        if not config.base_url or not config.api_key or not config.model:
            return JSONResponse({"error": "ai_config_required"}, status_code=status.HTTP_400_BAD_REQUEST)
        ai_client = OpenAICompatibleClient(
            base_url=config.base_url,
            api_key=config.api_key,
            model=config.model,
        )
    records = AttachmentProcessingService().process(campaign_id, attachments, ai_client=ai_client)
    request.app.state.database.save_extraction_records(records)
    return JSONResponse(
        {
            "campaign_id": campaign_id,
            "processed_count": len(records),
            "parsed_count": len([record for record in records if record.status == "parsed"]),
            "ai_processed_count": len([record for record in records if record.status == "ai_processed"]),
            "needs_ai_count": len([record for record in records if record.status == "needs_ai"]),
            "failed_count": len([record for record in records if record.status == "failed"]),
        }
    )


@router.post("/api/campaigns/{campaign_id}/send")
def send_campaign(campaign_id: str, request: Request) -> JSONResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    if _is_archived(campaign):
        return _archived_response()
    config = request.app.state.settings_store.load()
    recipients = request.app.state.database.list_recipients(campaign_id)
    attachments = request.app.state.database.list_campaign_attachments(campaign_id)
    try:
        result = CampaignSendService().send_campaign(
            campaign,
            recipients,
            config.smtp,
            attachments=[attachment.path for attachment in attachments],
        )
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=status.HTTP_400_BAD_REQUEST)

    request.app.state.database.save_recipients(recipients)
    for sent_message in result.sent_messages:
        request.app.state.database.save_sent_message(sent_message)
    campaign.status = "tracking" if result.sent_count else campaign.status
    request.app.state.database.save_campaign(campaign)
    return JSONResponse(
        {
            "sent_count": result.sent_count,
            "failed_count": result.failed_count,
            "campaign": _campaign_payload(campaign),
            "recipients": [_recipient_payload(recipient) for recipient in recipients],
        }
    )


@router.post("/api/campaigns/{campaign_id}/reminders/send")
def send_campaign_reminders(campaign_id: str, request: Request) -> JSONResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    if _is_archived(campaign):
        return _archived_response()
    config = request.app.state.settings_store.load()
    recipients = request.app.state.database.list_recipients(campaign_id)
    try:
        result = CampaignSendService().send_reminders(
            campaign,
            recipients,
            config.smtp,
            now=datetime.now(timezone.utc),
        )
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=status.HTTP_400_BAD_REQUEST)
    request.app.state.database.save_recipients(recipients)
    for sent_message in result.sent_messages:
        request.app.state.database.save_sent_message(sent_message)
    return JSONResponse(
        {
            "campaign_id": campaign_id,
            "sent_count": result.sent_count,
            "failed_count": result.failed_count,
            "recipients": [_recipient_payload(recipient) for recipient in recipients],
        }
    )


@router.post("/api/campaigns/{campaign_id}/export")
def export_campaign(campaign_id: str, request: Request) -> JSONResponse:
    campaign = request.app.state.database.get_campaign(campaign_id)
    if campaign is None:
        return JSONResponse({"error": "campaign_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    recipients = request.app.state.database.list_recipients(campaign_id)
    received_messages = request.app.state.database.list_received_messages(campaign_id)
    received_attachments = request.app.state.database.list_received_attachments(campaign_id)
    extraction_records = request.app.state.database.list_extraction_records(campaign_id)
    output_dir = request.app.state.settings.data_dir / "exports" / campaign_id
    campaign_dir = request.app.state.settings.data_dir / "campaigns" / campaign_id
    campaign_dir.mkdir(parents=True, exist_ok=True)
    exporter = ExportService()
    summary = exporter.write_summary(
        output_dir,
        campaign,
        recipients,
        received_messages=received_messages,
        received_attachments=received_attachments,
        extraction_records=extraction_records,
    )
    package_files = [message.body_path for message in received_messages]
    package_files.extend(attachment.path for attachment in received_attachments)
    package = exporter.write_package_files(output_dir, campaign_dir, summary, package_files)
    return JSONResponse({"summary": str(summary), "package": str(package)})


def _campaign_payload(campaign: Campaign) -> dict:
    return {
        "id": campaign.id,
        "name": campaign.name,
        "subject": campaign.subject,
        "deadline": campaign.deadline.isoformat(),
        "created_at": campaign.created_at.isoformat(),
        "archived_at": campaign.archived_at.isoformat() if campaign.archived_at else None,
        "status_before_archive": campaign.status_before_archive,
        "status": campaign.status,
        "reminder_strategy": campaign.reminder_strategy.value,
        "attachment_ai_enabled": campaign.attachment_ai_enabled,
    }


def _recipient_payload(recipient: CampaignRecipient) -> dict:
    return {
        "id": recipient.id,
        "email": recipient.email,
        "name": recipient.name,
        "company": recipient.company,
        "status": recipient.status.value,
        "sent_at": recipient.sent_at.isoformat() if recipient.sent_at else None,
        "replied_at": recipient.replied_at.isoformat() if recipient.replied_at else None,
        "reminded_at": recipient.reminded_at.isoformat() if recipient.reminded_at else None,
        "error": recipient.error,
        "excluded_from_required_reply": recipient.excluded_from_required_reply,
    }


def _received_message_payload(message) -> dict:
    return {
        "id": message.id,
        "from_email": message.from_email,
        "subject": message.subject,
        "message_id": message.message_id,
        "confidence": message.confidence,
        "body_summary": message.body_summary,
        "body_path": str(message.body_path),
        "needs_review": message.needs_review,
    }


def _is_archived(campaign: Campaign) -> bool:
    return campaign.status == "archived"


def _archived_response() -> JSONResponse:
    return JSONResponse({"error": "campaign_archived"}, status_code=status.HTTP_409_CONFLICT)


def _find_received_attachment(request: Request, campaign_id: str, attachment_id: str):
    return next(
        (
            item
            for item in request.app.state.database.list_received_attachments(campaign_id)
            if item.id == attachment_id
        ),
        None,
    )


def _find_campaign_attachment(request: Request, campaign_id: str, attachment_id: str):
    return next(
        (
            item
            for item in request.app.state.database.list_campaign_attachments(campaign_id)
            if item.id == attachment_id
        ),
        None,
    )


def _safe_local_file(request: Request, path: Path) -> Path | None:
    data_root = request.app.state.settings.data_dir.resolve()
    candidate = Path(path).resolve()
    if not candidate.is_relative_to(data_root):
        return None
    return candidate


def _download_file_response(
    request: Request,
    path: Path,
    filename: str,
    content_type: str,
):
    allowed_path = _safe_local_file(request, path)
    if allowed_path is None:
        return JSONResponse({"error": "file_not_allowed"}, status_code=status.HTTP_403_FORBIDDEN)
    if not allowed_path.exists():
        return JSONResponse({"error": "file_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    media_type = _effective_content_type(allowed_path, content_type or mimetypes.guess_type(filename)[0])
    return FileResponse(allowed_path, media_type=media_type, filename=filename)


def _inline_file_response(
    request: Request,
    path: Path,
    filename: str,
    content_type: str,
):
    allowed_path = _safe_local_file(request, path)
    if allowed_path is None:
        return JSONResponse({"error": "file_not_allowed"}, status_code=status.HTTP_403_FORBIDDEN)
    if not allowed_path.exists():
        return JSONResponse({"error": "file_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    media_type = _effective_content_type(allowed_path, content_type or mimetypes.guess_type(filename)[0])
    return FileResponse(allowed_path, media_type=media_type)


def _preview_file_response(request: Request, path: Path, content_type: str) -> JSONResponse:
    allowed_path = _safe_local_file(request, path)
    if allowed_path is None:
        return JSONResponse({"error": "file_not_allowed"}, status_code=status.HTTP_403_FORBIDDEN)
    if not allowed_path.exists():
        return JSONResponse({"error": "file_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    guessed_type = _effective_content_type(allowed_path, content_type)
    if allowed_path.suffix.lower() == ".xlsx":
        return _xlsx_preview_response(allowed_path, guessed_type)
    if _is_text_previewable(allowed_path, guessed_type):
        return JSONResponse(
            {
                "kind": "text",
                "filename": allowed_path.name,
                "content_type": guessed_type,
                "text": allowed_path.read_text(encoding="utf-8", errors="replace")[:20000],
            }
        )
    if guessed_type.startswith("image/") or guessed_type == "application/pdf":
        return JSONResponse(
            {
                "kind": "browser",
                "filename": allowed_path.name,
                "content_type": guessed_type,
                "message": "可在浏览器中预览，也可以用本机默认应用打开。",
            }
        )
    return JSONResponse(
        {
            "kind": "unsupported",
            "filename": allowed_path.name,
            "content_type": guessed_type,
            "message": "该文件类型暂不支持内嵌预览，请用本机默认应用打开。",
        }
    )


def _effective_content_type(path: Path, content_type: str | None) -> str:
    guessed = mimetypes.guess_type(path.name)[0]
    if not content_type or content_type in {"application/octet-stream", "binary/octet-stream"}:
        return guessed or "application/octet-stream"
    return content_type


def _xlsx_preview_response(path: Path, content_type: str) -> JSONResponse:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(max_row=50, max_col=12, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        cells = ["" if cell is None else str(cell) for cell in row]
        while cells and not cells[-1]:
            cells.pop()
        rows.append(cells)
    return JSONResponse(
        {
            "kind": "table",
            "filename": path.name,
            "content_type": content_type,
            "rows": rows,
        }
    )


def _is_text_previewable(path: Path, content_type: str) -> bool:
    return content_type.startswith("text/") or path.suffix.lower() in {
        ".csv",
        ".json",
        ".md",
        ".txt",
        ".log",
        ".tsv",
        ".xml",
        ".html",
    }


def _reveal_file_response(request: Request, path: Path) -> JSONResponse:
    allowed_path = _safe_local_file(request, path)
    if allowed_path is None:
        return JSONResponse({"error": "file_not_allowed"}, status_code=status.HTTP_403_FORBIDDEN)
    if not allowed_path.exists():
        return JSONResponse({"error": "file_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    if shutil.which("open") is None:
        return JSONResponse({"error": "finder_unavailable"}, status_code=status.HTTP_400_BAD_REQUEST)
    subprocess.run(["open", "-R", str(allowed_path)], check=False)
    return JSONResponse({"path": str(allowed_path), "revealed": True})


def _open_file_response(request: Request, path: Path) -> JSONResponse:
    allowed_path = _safe_local_file(request, path)
    if allowed_path is None:
        return JSONResponse({"error": "file_not_allowed"}, status_code=status.HTTP_403_FORBIDDEN)
    if not allowed_path.exists():
        return JSONResponse({"error": "file_not_found"}, status_code=status.HTTP_404_NOT_FOUND)
    if shutil.which("open") is None:
        return JSONResponse({"error": "file_open_unavailable"}, status_code=status.HTTP_400_BAD_REQUEST)
    subprocess.run(["open", str(allowed_path)], check=False)
    return JSONResponse({"path": str(allowed_path), "opened": True})


def _reminder_counts(request: Request, campaigns: list[Campaign]) -> dict[str, int]:
    service = ReminderService()
    now = datetime.now(timezone.utc)
    counts: dict[str, int] = {}
    for campaign in campaigns:
        recipients = request.app.state.database.list_recipients(campaign.id)
        counts[campaign.id] = len(service.due_recipients(campaign, recipients, now))
    return counts


def _sync_campaign_deadline_statuses(request: Request, campaigns: list[Campaign]) -> None:
    service = CampaignService()
    now = datetime.now(timezone.utc)
    for campaign in campaigns:
        recipients = request.app.state.database.list_recipients(campaign.id)
        before = (campaign.status, [recipient.status for recipient in recipients])
        service.refresh_campaign_status(campaign, recipients, now=now)
        after = (campaign.status, [recipient.status for recipient in recipients])
        if after != before:
            request.app.state.database.save_recipients(recipients)
            request.app.state.database.save_campaign(campaign)


def _prioritize_review_messages(
    messages,
    recipients: list[CampaignRecipient],
):
    recipient_emails = {recipient.email.lower() for recipient in recipients}
    return [message for message in messages if message.from_email.lower() in recipient_emails][:20]


def _create_campaign_objects(
    *,
    name: str,
    subject: str,
    body_template: str,
    deadline: datetime,
    recipients: list[RecipientInput],
    reminder_strategy: ReminderStrategy = ReminderStrategy.MANUAL_CONFIRM,
    attachment_ai_enabled: bool = False,
    cc: list[str] | None = None,
    bcc: list[str] | None = None,
) -> tuple[Campaign, list[CampaignRecipient]]:
    campaign = Campaign(
        id=uuid4().hex,
        name=name,
        subject=subject,
        body_template=body_template,
        deadline=deadline,
        reminder_strategy=reminder_strategy,
        status="draft",
        cc=cc or [],
        bcc=bcc or [],
        attachment_ai_enabled=attachment_ai_enabled,
    )
    campaign_recipients = [
        CampaignRecipient(
            id=uuid4().hex,
            campaign_id=campaign.id,
            email=recipient.email,
            name=recipient.name,
            company=recipient.company,
            status=RecipientStatus.DRAFT,
        )
        for recipient in recipients
    ]
    return campaign, campaign_recipients


def _new_recipients_for_campaign(
    campaign_id: str,
    rows: list[dict[str, str | None]],
    existing: list[CampaignRecipient],
) -> list[CampaignRecipient]:
    existing_emails = {recipient.email.lower() for recipient in existing}
    new_recipients: list[CampaignRecipient] = []
    for row in rows:
        email = (row.get("email") or "").strip()
        if not email or email.lower() in existing_emails:
            continue
        existing_emails.add(email.lower())
        new_recipients.append(
            CampaignRecipient(
                id=uuid4().hex,
                campaign_id=campaign_id,
                email=email,
                company=row.get("company"),
                name=row.get("name"),
                status=RecipientStatus.DRAFT,
            )
        )
    return new_recipients


def _settings_from_payload(payload: AppSettingsInput, existing: AppLocalConfig) -> AppLocalConfig:
    smtp = _mail_config(payload.smtp)
    imap = _mail_config(payload.imap)
    ai = _ai_config(payload.ai)
    if not smtp.password:
        smtp.password = existing.smtp.password
    if not imap.password:
        imap.password = existing.imap.password
    if not ai.api_key:
        ai.api_key = existing.ai.api_key
    return AppLocalConfig(smtp=smtp, imap=imap, ai=ai)


def _mail_config(payload: MailServerInput) -> MailServerConfig:
    return MailServerConfig.from_dict(payload.model_dump())


def _ai_config(payload: AiProviderInput) -> AiProviderConfig:
    return AiProviderConfig.from_dict(payload.model_dump())


def _split_addresses(value: str) -> list[str]:
    return [item.strip() for item in value.replace("\n", ",").split(",") if item.strip()]


def _parse_recipients_text(value: str) -> list[dict[str, str | None]]:
    recipients: list[dict[str, str | None]] = []
    for line in value.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        parts = [part.strip() for part in stripped.split(",")]
        if len(parts) == 1:
            recipients.append({"email": parts[0], "company": None, "name": None})
        elif len(parts) == 2:
            recipients.append({"email": parts[0], "company": parts[1], "name": None})
        else:
            recipients.append({"email": parts[0], "company": parts[1], "name": parts[2]})
    return recipients


def _parse_recipients_file(upload: UploadFile | None) -> list[dict[str, str | None]]:
    if upload is None or not upload.filename:
        return []
    content = upload.file.read()
    if not content:
        return []
    filename = upload.filename.lower()
    if filename.endswith(".xlsx"):
        return _parse_recipients_xlsx(content)
    return _parse_recipients_csv(content)


def _parse_recipients_csv(content: bytes) -> list[dict[str, str | None]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(StringIO(text))
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        return []
    if _looks_like_header(rows[0]):
        header = [cell.strip().lower() for cell in rows[0]]
        return [_recipient_from_mapping(dict(zip(header, row, strict=False))) for row in rows[1:]]
    return [_recipient_from_cells(row) for row in rows]


def _parse_recipients_xlsx(content: bytes) -> list[dict[str, str | None]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [
        ["" if cell is None else str(cell).strip() for cell in row]
        for row in sheet.iter_rows(values_only=True)
        if any(cell is not None and str(cell).strip() for cell in row)
    ]
    if not rows:
        return []
    if _looks_like_header(rows[0]):
        header = [cell.strip().lower() for cell in rows[0]]
        return [_recipient_from_mapping(dict(zip(header, row, strict=False))) for row in rows[1:]]
    return [_recipient_from_cells(row) for row in rows]


def _looks_like_header(row: list[str]) -> bool:
    normalized = {cell.strip().lower() for cell in row}
    return bool(normalized & {"email", "邮箱", "mail", "e-mail"})


def _recipient_from_mapping(row: dict[str, str]) -> dict[str, str | None]:
    email = _first_value(row, ["email", "邮箱", "mail", "e-mail"])
    company = _first_value(row, ["company", "公司", "单位", "organization"])
    name = _first_value(row, ["name", "姓名", "联系人", "contact"])
    return {"email": email, "company": company, "name": name}


def _first_value(row: dict[str, str], keys: list[str]) -> str | None:
    for key in keys:
        value = row.get(key)
        if value and str(value).strip():
            return str(value).strip()
    return None


def _recipient_from_cells(row: list[str]) -> dict[str, str | None]:
    cells = [str(cell).strip() for cell in row]
    return {
        "email": cells[0] if len(cells) > 0 else "",
        "company": cells[1] if len(cells) > 1 and cells[1] else None,
        "name": cells[2] if len(cells) > 2 and cells[2] else None,
    }


def _dedupe_recipients(rows: list[dict[str, str | None]]) -> list[dict[str, str | None]]:
    deduped: dict[str, dict[str, str | None]] = {}
    for row in rows:
        email = (row.get("email") or "").strip()
        if not email:
            continue
        key = email.lower()
        if key not in deduped:
            deduped[key] = {
                "email": email,
                "company": row.get("company"),
                "name": row.get("name"),
            }
    return list(deduped.values())


def _save_uploaded_attachments(
    request: Request,
    campaign_id: str,
    uploads: list[UploadFile],
) -> list[CampaignAttachment]:
    storage = CampaignStorage(request.app.state.settings.data_dir)
    saved: list[CampaignAttachment] = []
    for upload in uploads:
        if not upload.filename:
            continue
        content = upload.file.read()
        if not content:
            continue
        path = storage.save_outgoing_attachment(campaign_id, upload.filename, content)
        saved.append(
            CampaignAttachment(
                id=uuid4().hex,
                campaign_id=campaign_id,
                filename=upload.filename,
                path=path,
                content_type=upload.content_type or "application/octet-stream",
            )
        )
    return saved
