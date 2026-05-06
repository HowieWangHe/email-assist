from __future__ import annotations

import csv
import base64
import json
import mimetypes
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from app.models import ExtractionResult


class AttachmentExtractor:
    def extract_local(self, path: Path | str) -> ExtractionResult:
        file_path = Path(path)
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            with file_path.open(newline="", encoding="utf-8-sig") as handle:
                return ExtractionResult(status="parsed", rows=list(csv.DictReader(handle)))
        if suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return ExtractionResult(status="parsed", rows=[])
            headers = [str(value or "") for value in rows[0]]
            parsed = [
                {headers[index]: cell for index, cell in enumerate(row) if index < len(headers)}
                for row in rows[1:]
            ]
            return ExtractionResult(status="parsed", rows=parsed)
        if suffix in {".txt", ".md"}:
            return ExtractionResult(status="parsed", text=file_path.read_text(encoding="utf-8", errors="replace"))
        if suffix in {".pdf", ".doc", ".docx"}:
            return ExtractionResult(status="needs_ai", message="Text extraction and AI processing required")
        if suffix in {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"}:
            return ExtractionResult(status="needs_ai", message="OCR or vision AI processing required")
        return ExtractionResult(status="unsupported", message=f"Unsupported attachment type: {suffix or 'none'}")


class OpenAICompatibleClient:
    def __init__(
        self,
        *,
        base_url: str,
        api_key: str,
        model: str,
        http_client: httpx.Client | None = None,
    ):
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.model = model
        self.http_client = http_client or httpx.Client(timeout=60)

    def extract(self, instruction: str, content: str) -> dict[str, Any]:
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": instruction},
                {"role": "user", "content": content},
            ],
            "response_format": {"type": "json_object"},
        }
        return self._post_chat(payload)

    def extract_file(self, instruction: str, path: Path | str, content_type: str | None = None) -> dict[str, Any]:
        file_path = Path(path)
        mime_type = content_type or mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        encoded = base64.b64encode(file_path.read_bytes()).decode("ascii")
        if mime_type.startswith("image/"):
            user_content: str | list[dict[str, Any]] = [
                {
                    "type": "text",
                    "text": (
                        "Extract structured data from this inquiry reply attachment. "
                        "Return concise JSON with summary, detected fields, and any tabular data."
                    ),
                },
                {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{encoded}"}},
            ]
        else:
            user_content = (
                f"Filename: {file_path.name}\n"
                f"Content-Type: {mime_type}\n"
                f"Base64 attachment content:\n{encoded}"
            )
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": instruction},
                {"role": "user", "content": user_content},
            ],
            "response_format": {"type": "json_object"},
        }
        return self._post_chat(payload)

    def _post_chat(self, payload: dict[str, Any]) -> dict[str, Any]:
        response = self.http_client.post(
            f"{self.base_url}/chat/completions",
            headers={"Authorization": f"Bearer {self.api_key}"},
            json=payload,
        )
        response.raise_for_status()
        payload = response.json()
        content_text = payload["choices"][0]["message"]["content"]
        return json.loads(content_text)
