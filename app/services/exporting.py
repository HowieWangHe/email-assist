from __future__ import annotations

from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

from app.models import Campaign, CampaignRecipient, ExtractionRecord, ReceivedAttachment, ReceivedMessage


class ExportService:
    def write_summary(
        self,
        output_dir: Path | str,
        campaign: Campaign,
        recipients: list[CampaignRecipient],
        received_messages: list[ReceivedMessage] | None = None,
        received_attachments: list[ReceivedAttachment] | None = None,
        extraction_records: list[ExtractionRecord] | None = None,
    ) -> Path:
        received_messages = received_messages or []
        received_attachments = received_attachments or []
        extraction_records = extraction_records or []
        recipients_by_id = {recipient.id: recipient for recipient in recipients}
        target_dir = Path(output_dir)
        target_dir.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        info = workbook.active
        info.title = "campaign"
        info.append(["id", campaign.id])
        info.append(["name", campaign.name])
        info.append(["subject", campaign.subject])
        info.append(["status", campaign.status])
        info.append(["deadline", campaign.deadline.isoformat()])

        recipient_sheet = workbook.create_sheet("recipients")
        recipient_sheet.append(["email", "company", "status", "sent_at", "replied_at", "error"])
        for recipient in recipients:
            recipient_sheet.append(
                [
                    recipient.email,
                    recipient.company or "",
                    recipient.status.value,
                    recipient.sent_at.isoformat() if recipient.sent_at else "",
                    recipient.replied_at.isoformat() if recipient.replied_at else "",
                    recipient.error or "",
                ]
            )

        message_sheet = workbook.create_sheet("received_messages")
        message_sheet.append(
            ["recipient_email", "from_email", "subject", "message_id", "confidence", "body_summary", "body_path"]
        )
        for message in received_messages:
            recipient = recipients_by_id.get(message.recipient_id or "")
            message_sheet.append(
                [
                    recipient.email if recipient else "",
                    message.from_email,
                    message.subject,
                    message.message_id,
                    message.confidence,
                    message.body_summary,
                    str(message.body_path),
                ]
            )

        attachment_sheet = workbook.create_sheet("received_attachments")
        attachment_sheet.append(["recipient_email", "filename", "path", "content_type", "message_id"])
        for attachment in received_attachments:
            recipient = recipients_by_id.get(attachment.recipient_id or "")
            attachment_sheet.append(
                [
                    recipient.email if recipient else "",
                    attachment.filename,
                    str(attachment.path),
                    attachment.content_type,
                    attachment.message_id,
                ]
            )

        extraction_sheet = workbook.create_sheet("extraction_results")
        extraction_sheet.append(["filename", "status", "error", "result_json"])
        for record in extraction_records:
            extraction_sheet.append([record.filename, record.status, record.error, record.result_json])

        output = target_dir / "campaign_summary.xlsx"
        workbook.save(output)
        return output

    def write_package(self, output_dir: Path | str, campaign_dir: Path | str, excel_path: Path | str) -> Path:
        return self.write_package_files(output_dir, campaign_dir, excel_path)

    def write_package_files(
        self,
        output_dir: Path | str,
        campaign_dir: Path | str,
        excel_path: Path | str,
        files: list[Path] | None = None,
    ) -> Path:
        output_dir_path = Path(output_dir)
        output_dir_path.mkdir(parents=True, exist_ok=True)
        output = output_dir_path / "campaign_package.zip"
        campaign_root = Path(campaign_dir)
        export_root = output_dir_path.resolve()
        with ZipFile(output, "w", ZIP_DEFLATED) as archive:
            archive.write(excel_path, Path(excel_path).name)
            source_paths = files if files is not None else list(campaign_root.rglob("*"))
            for path in source_paths:
                path = Path(path)
                if not path.is_file():
                    continue
                try:
                    relative_path = path.relative_to(campaign_root)
                except ValueError:
                    relative_path = Path(path.name)
                if relative_path.parts and relative_path.parts[0] == "exports":
                    continue
                if path.resolve().is_relative_to(export_root):
                    continue
                archive.write(path, relative_path)
        return output
