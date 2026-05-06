from datetime import datetime, timedelta, timezone

from zipfile import ZipFile

from openpyxl import load_workbook

from app.models import (
    Campaign,
    CampaignRecipient,
    ExtractionRecord,
    ReceivedAttachment,
    ReceivedMessage,
    ReminderStrategy,
    RecipientStatus,
)
from app.services.exporting import ExportService
from app.services.extraction import AttachmentExtractor
from app.services.processing import AttachmentProcessingService
from app.services.storage import CampaignStorage


def test_archives_reply_body_and_attachment_with_safe_names(tmp_path):
    storage = CampaignStorage(tmp_path)

    paths = storage.archive_reply(
        campaign_id="c1",
        recipient_id="vendor/a",
        message_id="<reply@example.com>",
        body="hello",
        raw_email="raw",
        attachments={"报价/表.csv": b"name,price\nA,1\n"},
    )

    assert paths.body_path.read_text() == "hello"
    assert paths.raw_path.read_text() == "raw"
    assert paths.attachment_paths[0].name.endswith(".csv")
    assert "/" not in paths.attachment_paths[0].name


def test_saves_outgoing_campaign_attachment(tmp_path):
    storage = CampaignStorage(tmp_path)

    saved = storage.save_outgoing_attachment("c1", "报价/模板.xlsx", b"content")

    assert saved.read_bytes() == b"content"
    assert saved.parent.name == "original_attachments"
    assert "/" not in saved.name


def test_extracts_csv_rows_and_reports_unsupported_files(tmp_path):
    csv_path = tmp_path / "quote.csv"
    csv_path.write_text("item,price\nA,10\n")
    png_path = tmp_path / "image.png"
    png_path.write_bytes(b"not-real-image")

    extractor = AttachmentExtractor()

    csv_result = extractor.extract_local(csv_path)
    png_result = extractor.extract_local(png_path)

    assert csv_result.status == "parsed"
    assert csv_result.rows == [{"item": "A", "price": "10"}]
    assert png_result.status == "needs_ai"
    assert "OCR" in png_result.message


def test_extracts_plain_text_attachment(tmp_path):
    path = tmp_path / "reply.txt"
    path.write_text("quoted delivery date: Friday")

    result = AttachmentExtractor().extract_local(path)

    assert result.status == "parsed"
    assert result.text == "quoted delivery date: Friday"


def test_processes_received_attachments_into_records(tmp_path):
    path = tmp_path / "quote.csv"
    path.write_text("item,price\nA,10\n")
    attachment = ReceivedAttachment(
        id="a1",
        campaign_id="c1",
        recipient_id="r1",
        message_id="<reply@example.com>",
        filename="quote.csv",
        path=path,
        content_type="text/csv",
    )

    records = AttachmentProcessingService().process("c1", [attachment])

    assert records[0].attachment_id == "a1"
    assert records[0].status == "parsed"
    assert "\"price\": \"10\"" in records[0].result_json


def test_processes_needs_ai_attachment_with_ai_client(tmp_path):
    path = tmp_path / "reply.jpg"
    path.write_bytes(b"image-bytes")
    attachment = ReceivedAttachment(
        id="a1",
        campaign_id="c1",
        recipient_id="r1",
        message_id="<reply@example.com>",
        filename="reply.jpg",
        path=path,
        content_type="image/jpeg",
    )

    class FakeAI:
        def extract_file(self, instruction, attachment_path, content_type):
            assert "procurement" in instruction
            assert attachment_path == path
            assert content_type == "image/jpeg"
            return {"summary": "quote image"}

    records = AttachmentProcessingService().process("c1", [attachment], ai_client=FakeAI())

    assert records[0].status == "ai_processed"
    assert "quote image" in records[0].result_json


def test_exports_excel_summary_and_zip_package(tmp_path):
    campaign = Campaign(
        id="c1",
        name="May inquiry",
        subject="RFQ",
        body_template="Please quote.",
        deadline=datetime.now(timezone.utc) + timedelta(days=1),
        reminder_strategy=ReminderStrategy.MANUAL_CONFIRM,
        status="all_replied",
    )
    recipients = [
        CampaignRecipient(
            id="r1",
            campaign_id="c1",
            email="a@example.com",
            company="Vendor A",
            status=RecipientStatus.REPLIED,
        )
    ]
    campaign_dir = tmp_path / "campaign"
    campaign_dir.mkdir()
    (campaign_dir / "note.txt").write_text("hello")

    body_path = campaign_dir / "body.txt"
    raw_path = campaign_dir / "raw.eml"
    attachment_path = campaign_dir / "quote.csv"
    body_path.write_text("quote reply")
    raw_path.write_text("raw")
    attachment_path.write_text("item,price\nA,10\n")
    received_messages = [
        ReceivedMessage(
            id="m1",
            campaign_id="c1",
            recipient_id="r1",
            from_email="a@example.com",
            subject="Re: RFQ",
            message_id="<reply@example.com>",
            body_path=body_path,
            raw_path=raw_path,
            confidence="high",
            body_summary="quote reply summary",
        )
    ]
    received_attachments = [
        ReceivedAttachment(
            id="a1",
            campaign_id="c1",
            recipient_id="r1",
            message_id="<reply@example.com>",
            filename="quote.csv",
            path=attachment_path,
            content_type="text/csv",
        )
    ]
    extraction_records = [
        ExtractionRecord(
            id="x1",
            campaign_id="c1",
            attachment_id="a1",
            filename="quote.csv",
            status="parsed",
            result_json="{\"rows\": [{\"item\": \"A\"}]}",
        )
    ]

    export_dir = campaign_dir / "exports"
    export_dir.mkdir()
    (export_dir / "old_package.zip").write_text("old export")
    excel_path = ExportService().write_summary(
        tmp_path / "app_exports" / "c1",
        campaign,
        recipients,
        received_messages=received_messages,
        received_attachments=received_attachments,
        extraction_records=extraction_records,
    )
    zip_path = ExportService().write_package(tmp_path / "app_exports" / "c1", campaign_dir, excel_path)

    workbook = load_workbook(excel_path)

    assert "recipients" in workbook.sheetnames
    assert "received_messages" in workbook.sheetnames
    assert "received_attachments" in workbook.sheetnames
    assert "extraction_results" in workbook.sheetnames
    assert workbook["received_messages"]["F2"].value == "quote reply summary"
    assert workbook["received_attachments"]["A2"].value == "a@example.com"
    assert workbook["extraction_results"]["B2"].value == "parsed"
    assert zip_path.exists()
    assert zip_path.suffix == ".zip"
    with ZipFile(zip_path) as archive:
        names = archive.namelist()
    assert "exports/campaign_package.zip" not in names
    assert "exports/campaign_summary.xlsx" not in names
    assert "exports/old_package.zip" not in names


def test_exports_selected_files_without_scanning_campaign_directory(tmp_path):
    campaign_dir = tmp_path / "campaign"
    campaign_dir.mkdir()
    selected = campaign_dir / "recipients" / "r1" / "replies" / "m1" / "attachments" / "quote.csv"
    selected.parent.mkdir(parents=True)
    selected.write_text("item,price\nA,10\n")
    excluded = campaign_dir / "exports" / "old_package.zip"
    excluded.parent.mkdir()
    excluded.write_text("old export")
    unrelated = campaign_dir / "recipients" / "unmatched" / "raw.eml"
    unrelated.parent.mkdir(parents=True)
    unrelated.write_text("unmatched")
    summary = tmp_path / "campaign_summary.xlsx"
    summary.write_text("summary")

    package = ExportService().write_package_files(
        tmp_path / "exports" / "c1",
        campaign_dir,
        summary,
        files=[selected],
    )

    with ZipFile(package) as archive:
        names = archive.namelist()

    assert "campaign_summary.xlsx" in names
    assert "recipients/r1/replies/m1/attachments/quote.csv" in names
    assert "exports/old_package.zip" not in names
    assert "recipients/unmatched/raw.eml" not in names
